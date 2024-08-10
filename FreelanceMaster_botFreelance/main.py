import telebot
from openpyxl import load_workbook
from io import BytesIO
from telebot import types
import pandas as pd
from datetime import datetime

TOKEN = 'YOUR_BOT_TOKEN'
bot = telebot.TeleBot(TOKEN)

FILE_PATH = r"D:\Фриланс\ЗП\ЗП - ФРИЛАНС.xlsx"
selected_sheet = None
user_data = {}

def send_message(chat_id, text, reply_markup=None):
    try:
        bot.send_message(chat_id, text, reply_markup=reply_markup)
    except Exception as e:
        print(f"Error sending message: {e}")

def create_main_menu_markup():
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    buttons = ["📄 Выбрать лист", "✏️ Редактировать данные", "➕ Добавить данные", "💰 Рассчитать сумму", "🗑️ Удалить данные"]
    markup.add(*[types.KeyboardButton(btn) for btn in buttons])
    return markup

def create_action_menu_markup():
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    buttons = ["✏️ Редактировать данные", "➕ Добавить данные", "💰 Рассчитать сумму", "🗑️ Удалить данные", "🏠 Назад в главное меню"]
    markup.add(*[types.KeyboardButton(btn) for btn in buttons])
    return markup

@bot.message_handler(commands=['start'])
def send_welcome(message):
    markup = create_main_menu_markup()
    send_message(message.chat.id, "Добро пожаловать! Выберите действие:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "📄 Выбрать лист")
def choose_sheet(message):
    sheets = get_sheet_names()
    if not sheets:
        send_message(message.chat.id, "Не удалось получить список листов.")
        return

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add(*[types.KeyboardButton(sheet) for sheet in sheets])
    send_message(message.chat.id, "Выберите лист для редактирования:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text in get_sheet_names())
def set_sheet(message):
    global selected_sheet
    selected_sheet = message.text
    user_data[message.chat.id] = {"step": "none"}
    markup = create_action_menu_markup()
    send_message(message.chat.id, f"Выбран лист: {selected_sheet}. Выберите действие:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "✏️ Редактировать данные")
def start_edit_data(message):
    if not selected_sheet:
        send_message(message.chat.id, "Сначала выберите лист.")
        return

    user_data[message.chat.id] = {"step": "edit_id"}
    send_message(message.chat.id, "Введите ID записи для редактирования:")

@bot.message_handler(func=lambda message: user_data.get(message.chat.id, {}).get("step") == "edit_id")
def get_edit_id(message):
    user_data[message.chat.id]["edit_id"] = message.text.strip()
    user_data[message.chat.id]["step"] = "edit_data"
    send_message(message.chat.id, "Введите новые значения в формате: 'Сумма грн, Выполнил, Дата выплаты, Пришли деньги'. Пример: '700, ✅, 13.06.2024, ✅'")

@bot.message_handler(func=lambda message: user_data.get(message.chat.id, {}).get("step") == "edit_data")
def process_edit_data(message):
    data = message.text.split(',', 3)

    if len(data) != 4:
        send_message(message.chat.id, "Неверный формат данных. Введите данные в формате: 'Сумма грн, Выполнил, Дата выплаты, Пришли деньги'.")
        return

    amount, completed, payout_date, received = map(str.strip, data)
    id_value = user_data[message.chat.id].get("edit_id")

    if not id_value.isdigit():
        send_message(message.chat.id, "ID должно быть числом.")
        return

    id_value = int(id_value)
    payout_date = parse_date(payout_date)

    if not payout_date:
        send_message(message.chat.id, "Дата в неверном формате. Используйте формат 'ДД.ММ.ГГГГ'.")
        return

    try:
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[selected_sheet]
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == id_value:
                row[1].value = amount
                row[2].value = completed
                row[3].value = format_date(payout_date)
                row[4].value = received
                workbook.save(FILE_PATH)
                send_message(message.chat.id, "Данные успешно обновлены!")
                user_data[message.chat.id]["step"] = "none"
                return
        send_message(message.chat.id, "ID не найден в таблице.")
    except Exception as e:
        send_message(message.chat.id, f"Ошибка при редактировании данных: {e}")
        user_data[message.chat.id]["step"] = "none"

@bot.message_handler(func=lambda message: message.text == "➕ Добавить данные")
def add_data(message):
    if not selected_sheet:
        send_message(message.chat.id, "Сначала выберите лист.")
        return
    send_message(message.chat.id, "Отправьте данные для добавления в формате: 'ID, Сумма грн, Выполнил, Дата выплаты, Пришли деньги'. Пример: '14569, 800, ✅, 13.06.2024, ❌'.")

@bot.message_handler(func=lambda message: message.text == "💰 Рассчитать сумму")
def calculate_sum(message):
    if not selected_sheet:
        send_message(message.chat.id, "Сначала выберите лист.")
        return

    try:
        df = pd.read_excel(FILE_PATH, sheet_name=selected_sheet)
        total_sum = df['Сумма грн'].sum()
        send_message(message.chat.id, f"Общая сумма: {total_sum:.2f} грн")
    except Exception as e:
        send_message(message.chat.id, f"Ошибка при расчете суммы: {e}")

@bot.message_handler(func=lambda message: message.text == "🗑️ Удалить данные")
def start_delete_data(message):
    if not selected_sheet:
        send_message(message.chat.id, "Сначала выберите лист.")
        return

    user_data[message.chat.id] = {"step": "delete_id"}
    send_message(message.chat.id, "Введите ID записи для удаления:")

@bot.message_handler(func=lambda message: user_data.get(message.chat.id, {}).get("step") == "delete_id")
def delete_data(message):
    id_value = message.text.strip()

    if not id_value.isdigit():
        send_message(message.chat.id, "ID должно быть числом.")
        return

    id_value = int(id_value)

    try:
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[selected_sheet]
        rows_to_delete = []

        for row in sheet.iter_rows(min_row=2):
            if row[0].value == id_value:
                rows_to_delete.append(row[0].row)

        if not rows_to_delete:
            send_message(message.chat.id, "ID не найден в таблице.")
            return

        for row_num in reversed(rows_to_delete):
            sheet.delete_rows(row_num)

        workbook.save(FILE_PATH)
        send_message(message.chat.id, "Данные успешно удалены!")
    except Exception as e:
        send_message(message.chat.id, f"Ошибка при удалении данных: {e}")

    user_data[message.chat.id]["step"] = "none"

@bot.message_handler(func=lambda message: ',' in message.text)
def process_data(message):
    parts = message.text.split(',', 4)

    if len(parts) == 5:  # Handling Add Data
        handle_add_data(message, parts)

def handle_add_data(message, parts):
    id_value, amount, completed, payout_date, received = map(str.strip, parts)
    id_value = int(id_value)
    amount = float(amount)
    payout_date = parse_date(payout_date)

    if not payout_date:
        send_message(message.chat.id, "Дата в неверном формате. Используйте формат 'ДД.ММ.ГГГГ'.")
        return

    payout_date = format_date(payout_date)

    try:
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[selected_sheet]
        row_exists = False
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == id_value:
                row_exists = True
                row[1].value = amount
                row[2].value = completed
                row[3].value = payout_date
                row[4].value = received
                break
        if not row_exists:
            sheet.append([id_value, amount, completed, payout_date, received])
        workbook.save(FILE_PATH)
        send_message(message.chat.id, "Данные успешно добавлены!")
    except Exception as e:
        send_message(message.chat.id, f"Ошибка при добавлении данных: {e}")

def parse_date(date_str):
    try:
        return datetime.strptime(date_str, "%d.%m.%Y")
    except ValueError:
        return None

def format_date(date_obj):
    return date_obj.strftime("%d.%m.%Y")

def get_sheet_names():
    try:
        workbook = load_workbook(FILE_PATH)
        return workbook.sheetnames
    except Exception as e:
        print(f"Error getting sheet names: {e}")
        return []

@bot.message_handler(func=lambda message: message.text == "🏠 Назад в главное меню")
def back_to_main_menu(message):
    user_data[message.chat.id] = {"step": "none"}
    markup = create_main_menu_markup()
    send_message(message.chat.id, "Вы вернулись в главное меню. Выберите действие:", reply_markup=markup)


if __name__ == "__main__":
    bot.polling(none_stop=True)
